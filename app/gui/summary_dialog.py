# gui/summary_dialog.py

import os
import tempfile
import re
import numpy as np
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QMessageBox, QFileDialog, QProgressBar, QCheckBox, QScrollArea, QWidget,
    QFrame, QTextEdit, QLineEdit, QGroupBox, QComboBox, QGridLayout
)
from PyQt6.QtCore import Qt, pyqtSignal
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
import logging
from scipy import interpolate

logger = logging.getLogger(__name__)


class GraphData:
    """Класс для хранения данных графика с поддержкой операций"""
    
    def __init__(self, freqs, amps, label=""):
        self.freqs = np.asarray(freqs)
        self.amps = np.asarray(amps)
        self.label = label
        self._validate_data()
    
    def _validate_data(self):
        """Проверка и очистка данных"""
        if len(self.freqs) == 0 or len(self.amps) == 0:
            raise ValueError("Пустые данные")
        
        if len(self.freqs) != len(self.amps):
            raise ValueError("Длины частот и амплитуд не совпадают")
        
        # Убираем NaN и inf
        valid_mask = np.isfinite(self.freqs) & np.isfinite(self.amps)
        if not np.any(valid_mask):
            raise ValueError("Нет валидных данных")
        
        self.freqs = self.freqs[valid_mask]
        self.amps = self.amps[valid_mask]
        
        # Сортируем по частоте
        sort_idx = np.argsort(self.freqs)
        self.freqs = self.freqs[sort_idx]
        self.amps = self.amps[sort_idx]
    
    def interpolate_to_freqs(self, target_freqs):
        """Интерполяция на заданные частоты"""
        if len(self.freqs) < 2:
            raise ValueError("Слишком мало точек для интерполяции")
        
        # Проверяем диапазон
        if target_freqs[0] < self.freqs[0] or target_freqs[-1] > self.freqs[-1]:
            logger.warning(f"Интерполяция за пределами диапазона: {target_freqs[0]}-{target_freqs[-1]} vs {self.freqs[0]}-{self.freqs[-1]}")
        
        # Линейная интерполяция
        interp_func = interpolate.interp1d(self.freqs, self.amps, 
                                          kind='linear', 
                                          bounds_error=False,
                                          fill_value=(self.amps[0], self.amps[-1]))
        return interp_func(target_freqs)
    
    @staticmethod
    def create_common_freq_grid(*graphs, step_ratio=0.1):
        """Создание общей частотной сетки для нескольких графиков"""
        all_freqs = []
        for graph in graphs:
            all_freqs.extend(graph.freqs)
        
        if not all_freqs:
            raise ValueError("Нет данных для создания сетки")
        
        all_freqs = np.array(all_freqs)
        freq_min = np.min(all_freqs)
        freq_max = np.max(all_freqs)
        
        # Определяем шаг как минимальный шаг среди всех графиков
        steps = []
        for graph in graphs:
            if len(graph.freqs) > 1:
                step = np.min(np.diff(graph.freqs))
                steps.append(step)
        
        if steps:
            step_size = np.min(steps) * step_ratio
        else:
            step_size = (freq_max - freq_min) / 1000
        
        # Создаем сетку
        common_freqs = np.arange(freq_min, freq_max + step_size, step_size)
        
        # Логируем параметры сетки
        logger.info(f"Общая сетка: {freq_min:.2f}-{freq_max:.2f} Гц, шаг {step_size:.4f}, точек: {len(common_freqs)}")
        
        return common_freqs
    
    def __add__(self, other):
        return self._operation(other, 'add')
    
    def __sub__(self, other):
        return self._operation(other, 'sub')
    
    def __mul__(self, other):
        return self._operation(other, 'mul')
    
    def __truediv__(self, other):
        return self._operation(other, 'div')
    
    def _operation(self, other, op):
        """Выполнение операции с другим графиком или числом"""
        if isinstance(other, (int, float, np.number)):
            # Операция с числом
            if op == 'add':
                new_amps = self.amps + other
            elif op == 'sub':
                new_amps = self.amps - other
            elif op == 'mul':
                new_amps = self.amps * other
            elif op == 'div':
                new_amps = self.amps / other
            else:
                raise ValueError(f"Неизвестная операция: {op}")
            
            return GraphData(self.freqs, new_amps, 
                           f"({self.label}) {op} {other}")
        
        elif isinstance(other, GraphData):
            # Операция с другим графиком
            # Создаем общую частотную сетку
            common_freqs = self.create_common_freq_grid(self, other)
            
            # Интерполируем оба графика на общую сетку
            amps1 = self.interpolate_to_freqs(common_freqs)
            amps2 = other.interpolate_to_freqs(common_freqs)
            
            # Выполняем операцию
            if op == 'add':
                new_amps = amps1 + amps2
            elif op == 'sub':
                new_amps = amps1 - amps2
            elif op == 'mul':
                new_amps = amps1 * amps2
            elif op == 'div':
                new_amps = amps1 / (amps2 + 1e-10)  # Избегаем деления на 0
            else:
                raise ValueError(f"Неизвестная операция: {op}")
            
            return GraphData(common_freqs, new_amps, 
                           f"({self.label}) {op} ({other.label})")
        else:
            raise TypeError(f"Неподдерживаемый тип: {type(other)}")


class ExpressionParser:
    """Парсер математических выражений для графиков"""
    
    def __init__(self, frequency_responses):
        self.frequency_responses = frequency_responses
        self.functions = {
            'MAX': self._max,
            'MIN': self._min,
            'MEAN': self._mean,
            'SUM': self._sum,
            'ABS': self._abs,
            'SQRT': self._sqrt,
            'LOG': self._log,
            'LOG10': self._log10,
            'INDEX': self._index,
            'AVG': self._mean,
            'STD': self._std,
        }
    
    def parse_expression(self, expression):
        """Парсинг математического выражения"""
        try:
            # Подготовка выражения
            expr = expression.strip()
            logger.info(f"Парсинг выражения: {expr}")
            
            # Извлекаем все идентификаторы графиков
            pattern = r'([A-Za-z0-9_]+)\[([A-Za-z0-9_]+)\](?:\[([^\]]+)\])?'
            matches = re.findall(pattern, expr)
            
            if not matches:
                # Пробуем вычислить как числовое выражение
                return float(eval(expr))
            
            # Заменяем идентификаторы на временные переменные
            var_mapping = {}
            for i, (graph_id, channel, slice_expr) in enumerate(matches):
                var_name = f'__var{i}__'
                var_mapping[var_name] = (graph_id, channel, slice_expr)
                
                # Заменяем в выражении
                old_str = f"{graph_id}[{channel}]"
                if slice_expr:
                    old_str += f"[{slice_expr}]"
                expr = expr.replace(old_str, var_name)
            
            # Компилируем выражение
            local_vars = {}
            local_vars.update(self.functions)
            
            # Вычисляем значения переменных
            for var_name, (graph_id, channel, slice_expr) in var_mapping.items():
                graph_data = self._get_graph_data(graph_id, channel, slice_expr)
                local_vars[var_name] = graph_data
            
            # Выполняем вычисление
            result = eval(expr, {"__builtins__": {}}, local_vars)
            
            # Обрабатываем результат
            if isinstance(result, GraphData):
                return result
            elif isinstance(result, (int, float, np.number)):
                return float(result)
            else:
                raise TypeError(f"Неизвестный тип результата: {type(result)}")
            
        except Exception as e:
            logger.error(f"Ошибка парсинга выражения '{expression}': {str(e)}", exc_info=True)
            raise
    
    def _get_graph_data(self, graph_id, channel, slice_expr=None):
        """Получение данных графика по идентификатору"""
        logger.info(f"Поиск графика: {graph_id}[{channel}], срез: {slice_expr}")
        
        # Ищем график в frequency_responses
        for key, (freqs, amps) in self.frequency_responses.items():
            subject_code, analysis_index, channel_name = key
            full_id = f"{subject_code}_{analysis_index}"
            
            if full_id == graph_id and channel_name == channel:
                logger.info(f"Найден график: {subject_code}_{analysis_index}_{channel_name}")
                
                if slice_expr:
                    freqs, amps = self._apply_slice(freqs, amps, slice_expr)
                
                return GraphData(freqs, amps, f"{graph_id}[{channel}]")
        
        # Если не нашли, пробуем найти по другому формату
        for key, (freqs, amps) in self.frequency_responses.items():
            subject_code, analysis_index, channel_name = key
            if subject_code == graph_id and channel_name == channel:
                logger.info(f"Найден график (альтернативный формат): {subject_code}_{analysis_index}_{channel_name}")
                
                if slice_expr:
                    freqs, amps = self._apply_slice(freqs, amps, slice_expr)
                
                return GraphData(freqs, amps, f"{graph_id}[{channel}]")
        
        # Выводим доступные графики для отладки
        available = []
        for key in self.frequency_responses.keys():
            subject_code, analysis_index, channel_name = key
            available.append(f"{subject_code}_{analysis_index}[{channel_name}]")
        
        logger.error(f"Доступные графики: {available}")
        raise ValueError(f"График {graph_id}[{channel}] не найден. Доступные: {', '.join(available)}")
    
    def _apply_slice(self, freqs, amps, slice_expr):
        """Применение среза к данным"""
        try:
            # Удаляем пробелы
            slice_expr = slice_expr.strip()
            
            if ':' in slice_expr:
                # Срез по частоте
                parts = slice_expr.split(':')
                if len(parts) == 2:
                    freq_start = float(parts[0]) if parts[0] else None
                    freq_end = float(parts[1]) if parts[1] else None
                    
                    mask = np.ones_like(freqs, dtype=bool)
                    if freq_start is not None:
                        mask = mask & (freqs >= freq_start)
                    if freq_end is not None:
                        mask = mask & (freqs <= freq_end)
                    
                    return freqs[mask], amps[mask]
            else:
                # Одиночная частота или индекс
                try:
                    # Пробуем как индекс
                    idx = int(slice_expr)
                    if 0 <= idx < len(freqs):
                        return np.array([freqs[idx]]), np.array([amps[idx]])
                except ValueError:
                    # Пробуем как частоту
                    target_freq = float(slice_expr)
                    # Находим ближайшую частоту
                    idx = np.argmin(np.abs(freqs - target_freq))
                    return np.array([freqs[idx]]), np.array([amps[idx]])
        
        except Exception as e:
            logger.error(f"Ошибка применения среза '{slice_expr}': {str(e)}")
            raise ValueError(f"Некорректный срез: {slice_expr}")
    
    def _max(self, data):
        """Максимальное значение"""
        if isinstance(data, GraphData):
            return np.max(data.amps)
        return np.max(data)
    
    def _min(self, data):
        """Минимальное значение"""
        if isinstance(data, GraphData):
            return np.min(data.amps)
        return np.min(data)
    
    def _mean(self, data):
        """Среднее значение"""
        if isinstance(data, GraphData):
            return np.mean(data.amps)
        return np.mean(data)
    
    def _std(self, data):
        """Стандартное отклонение"""
        if isinstance(data, GraphData):
            return np.std(data.amps)
        return np.std(data)
    
    def _sum(self, data):
        """Сумма значений"""
        if isinstance(data, GraphData):
            return np.sum(data.amps)
        return np.sum(data)
    
    def _abs(self, data):
        """Абсолютное значение"""
        if isinstance(data, GraphData):
            return GraphData(data.freqs, np.abs(data.amps), f"ABS({data.label})")
        return np.abs(data)
    
    def _sqrt(self, data):
        """Квадратный корень"""
        if isinstance(data, GraphData):
            return GraphData(data.freqs, np.sqrt(np.abs(data.amps)), f"SQRT({data.label})")
        return np.sqrt(np.abs(data))
    
    def _log(self, data):
        """Натуральный логарифм"""
        if isinstance(data, GraphData):
            return GraphData(data.freqs, np.log(np.abs(data.amps) + 1e-10), f"LOG({data.label})")
        return np.log(np.abs(data) + 1e-10)
    
    def _log10(self, data):
        """Десятичный логарифм"""
        if isinstance(data, GraphData):
            return GraphData(data.freqs, np.log10(np.abs(data.amps) + 1e-10), f"LOG10({data.label})")
        return np.log10(np.abs(data) + 1e-10)
    
    def _index(self, data, idx):
        """Значение по индексу"""
        if isinstance(data, GraphData):
            if 0 <= idx < len(data.amps):
                return data.amps[idx]
            else:
                raise IndexError(f"Индекс {idx} вне диапазона [0, {len(data.amps)})")
        return data[idx] if hasattr(data, '__getitem__') else data


class LegendWidget(QWidget):
    """Виджет легенды с чекбоксами"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.checkboxes = {}
        self.lines_mapping = {}  # checkbox -> line object
        self.setup_ui()

    def setup_ui(self):
        """Настройка интерфейса легенды"""
        layout = QVBoxLayout()

        # Заголовок
        title_label = QLabel('Легенда')
        title_label.setStyleSheet('font-weight: bold; margin: 5px;')
        layout.addWidget(title_label)

        # Фрейм для чекбоксов
        self.checkbox_frame = QFrame()
        self.checkbox_layout = QVBoxLayout(self.checkbox_frame)
        self.checkbox_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        # Область прокрутки
        scroll_area = QScrollArea()
        scroll_area.setWidget(self.checkbox_frame)
        scroll_area.setWidgetResizable(True)
        scroll_area.setMaximumWidth(250)
        scroll_area.setMinimumWidth(200)

        layout.addWidget(scroll_area)
        self.setLayout(layout)

    def add_line(self, label, line, color):
        """Добавление линии в легенду"""
        checkbox = QCheckBox(label)
        checkbox.setChecked(True)
        checkbox.setStyleSheet(f"QCheckBox {{ color: {color}; }}")

        # Сохраняем связь
        self.checkboxes[label] = checkbox
        self.lines_mapping[checkbox] = line

        self.checkbox_layout.addWidget(checkbox)

        # Подключаем сигнал
        checkbox.stateChanged.connect(self.on_checkbox_changed)

    def on_checkbox_changed(self):
        """Обработка изменения состояния чекбокса"""
        checkbox = self.sender()
        if checkbox in self.lines_mapping:
            line = self.lines_mapping[checkbox]
            visible = checkbox.isChecked()
            line.set_visible(visible)

            # Передаем сигнал родительскому виджету
            if hasattr(self.parent(), 'on_legend_visibility_changed'):
                self.parent().on_legend_visibility_changed()


class SummaryDialog(QDialog):
    """Диалог для построения сводного графика АЧХ из выбранных анализов"""

    def __init__(self, data_manager, tree_manager, parent=None):
        super().__init__(parent)
        self.data_manager = data_manager
        self.tree_manager = tree_manager
        self.setWindowTitle('Сводный график АЧХ')
        self.setGeometry(100, 100, 1600, 1000)

        # Данные для графиков
        self.frequency_responses = {}  # (subject_code, analysis_index, channel_name) -> (freqs, response)
        self.lines = {}  # (subject_code, analysis_index, channel_name) -> line object
        self.expression_lines = {}  # Идентификаторы графиков, созданных через выражения
        self.parser = None

        self.setup_ui()
        self.load_selected_analyses()

    def setup_ui(self):
        """Настройка интерфейса"""
        main_layout = QHBoxLayout()

        # Левая часть - график и управление
        left_layout = QVBoxLayout()

        # Заголовок
        title_label = QLabel('Сводный график АЧХ выбранных анализов')
        title_label.setStyleSheet('font-size: 14px; font-weight: bold; margin: 10px;')
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_layout.addWidget(title_label)

        # График
        self.setup_plot(left_layout)

        # Панель математических выражений
        self.setup_expression_panel(left_layout)

        # Элементы управления
        self.setup_controls(left_layout)

        # Кнопки
        self.setup_buttons(left_layout)

        # Правая часть - легенда
        self.legend_widget = LegendWidget(self)

        # Добавляем обе части в основной layout
        main_layout.addLayout(left_layout, 4)  # 4/5 ширины для графика
        main_layout.addWidget(self.legend_widget, 1)  # 1/5 ширины для легенды

        self.setLayout(main_layout)

    def setup_plot(self, layout):
        """Настройка области графика"""
        # Создаем фигуру и canvas
        self.figure = Figure(figsize=(10, 8))
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)

        layout.addWidget(self.toolbar)
        layout.addWidget(self.canvas)

        # Создаем оси
        self.ax = self.figure.add_subplot(111)
        self.ax.set_xlabel('Частота (Гц)')
        self.ax.set_ylabel('Амплитуда (В)')
        self.ax.set_title('Амплитудно-частотная характеристика')
        self.ax.grid(True, alpha=0.3)

        # Сохраняем исходные пределы осей
        self.original_xlim = None
        self.original_ylim = None

    def setup_expression_panel(self, layout):
        """Настройка панели для математических выражений"""
        group_box = QGroupBox("Математические выражения")
        group_layout = QVBoxLayout()
        
        # Поле ввода выражения
        expression_layout = QHBoxLayout()
        expression_layout.addWidget(QLabel("Выражение:"))
        
        self.expression_input = QLineEdit()
        self.expression_input.setPlaceholderText("AN2_1[CH1] * AN2_2[CH2] или MAX(AN2_1[CH1])")
        expression_layout.addWidget(self.expression_input)
        
        # Кнопка вычисления
        self.calc_btn = QPushButton("Вычислить")
        self.calc_btn.clicked.connect(self.calculate_expression)
        expression_layout.addWidget(self.calc_btn)
        
        group_layout.addLayout(expression_layout)
        
        # Поле результата
        result_layout = QHBoxLayout()
        result_layout.addWidget(QLabel("Результат:"))
        
        self.result_display = QTextEdit()
        self.result_display.setMaximumHeight(100)
        self.result_display.setReadOnly(True)
        result_layout.addWidget(self.result_display)
        
        group_layout.addLayout(result_layout)
        
        # Информация о синхронизации
        self.sync_info_label = QLabel("")
        self.sync_info_label.setStyleSheet("color: blue; font-style: italic;")
        group_layout.addWidget(self.sync_info_label)
        
        # Примеры выражений
        examples_label = QLabel("Примеры выражений (используйте AN2_1, AN2_2 и т.д.):")
        examples_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        group_layout.addWidget(examples_label)
        
        examples_text = QTextEdit()
        examples_text.setMaximumHeight(120)
        examples_text.setReadOnly(True)
        examples_text.setPlainText(
            "AN2_1[CH1] * AN2_2[CH2] - поэлементное умножение с синхронизацией\n"
            "AN2_1[CH1][10:100] / MAX(AN2_2[CH2]) - срез по частоте и деление\n"
            "AN2_1[CH1] + AN2_2[CH2] - сложение с синхронизацией\n"
            "AN2_1[CH1] - AN2_2[CH2] - вычитание\n"
            "SQRT(AN2_1[CH1]) - квадратный корень\n"
            "LOG10(AN2_1[CH1]) - десятичный логарифм\n"
            "MAX(AN2_1[CH1]) - максимальное значение\n"
            "INDEX(MAX(AN2_1[CH1]), 1) - индекс максимального значения\n"
            "AN2_1[CH1][50] - значение на частоте 50 Гц"
        )
        group_layout.addWidget(examples_text)
        
        # Кнопка добавления графика
        self.add_graph_btn = QPushButton("Добавить как новый график")
        self.add_graph_btn.clicked.connect(self.add_expression_graph)
        self.add_graph_btn.setEnabled(False)
        group_layout.addWidget(self.add_graph_btn)
        
        group_box.setLayout(group_layout)
        layout.addWidget(group_box)
        
        # Сохраняем последний результат
        self.last_calculation_result = None

    def setup_controls(self, layout):
        """Настройка элементов управления"""
        controls_layout = QHBoxLayout()

        # Прогресс-бар
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)

        # Чекбокс для автоматического обновления
        self.auto_update_cb = QCheckBox("Автообновление масштаба")
        self.auto_update_cb.setChecked(True)

        # Кнопка сброса масштаба
        self.reset_zoom_btn = QPushButton('Сбросить масштаб')
        self.reset_zoom_btn.clicked.connect(self.reset_zoom)

        controls_layout.addWidget(self.progress_bar)
        controls_layout.addWidget(self.auto_update_cb)
        controls_layout.addWidget(self.reset_zoom_btn)
        controls_layout.addStretch()

        layout.addLayout(controls_layout)

    def setup_buttons(self, layout):
        """Настройка кнопок"""
        buttons_layout = QHBoxLayout()

        # Кнопка обновления
        self.update_btn = QPushButton('Обновить график')
        self.update_btn.clicked.connect(self.load_selected_analyses)

        # Кнопка экспорта
        self.export_btn = QPushButton('Экспорт в Excel')
        self.export_btn.clicked.connect(self.export_to_excel)

        # Кнопка закрытия
        self.close_btn = QPushButton('Закрыть')
        self.close_btn.clicked.connect(self.close)

        buttons_layout.addWidget(self.update_btn)
        buttons_layout.addWidget(self.export_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.close_btn)

        layout.addLayout(buttons_layout)

    def calculate_expression(self):
        """Вычисление математического выражения"""
        expression = self.expression_input.text().strip()
        if not expression:
            QMessageBox.warning(self, "Предупреждение", "Введите выражение")
            return
        
        try:
            # Создаем парсер, если его еще нет
            if self.parser is None:
                self.parser = ExpressionParser(self.frequency_responses)
            
            # Вычисляем выражение
            result = self.parser.parse_expression(expression)
            
            # Очищаем информацию о синхронизации
            self.sync_info_label.setText("")
            
            # Отображаем результат
            if isinstance(result, GraphData):
                # Это график
                result_text = f"График: {result.label}\n"
                result_text += f"Точек: {len(result.freqs)}\n"
                result_text += f"Диапазон частот: {result.freqs[0]:.2f} - {result.freqs[-1]:.2f} Гц\n"
                result_text += f"Диапазон амплитуд: {np.min(result.amps):.4e} - {np.max(result.amps):.4e} В\n"
                result_text += f"Шаг частоты: {np.mean(np.diff(result.freqs)):.4f} Гц"
                
                # Показываем информацию о синхронизации, если это результат операции
                if ')' in result.label and '(' in result.label:
                    self.sync_info_label.setText("✓ Графики синхронизированы по общей частотной сетке")
                
                # Сохраняем результат для возможного добавления как график
                self.last_calculation_result = {
                    'type': 'graph',
                    'graph_data': result,
                    'expression': expression
                }
                self.add_graph_btn.setEnabled(True)
                
            elif isinstance(result, (int, float)):
                # Это скалярное значение
                result_text = f"Скалярный результат: {result:.6e}"
                
                # Сохраняем результат
                self.last_calculation_result = {
                    'type': 'scalar',
                    'value': result,
                    'expression': expression
                }
                self.add_graph_btn.setEnabled(False)
            else:
                result_text = f"Неизвестный тип результата: {type(result)}"
                self.last_calculation_result = None
                self.add_graph_btn.setEnabled(False)
            
            self.result_display.setPlainText(result_text)
            
        except Exception as e:
            error_msg = f"Ошибка вычисления выражения: {str(e)}"
            logger.error(error_msg)
            self.result_display.setPlainText(error_msg)
            self.last_calculation_result = None
            self.add_graph_btn.setEnabled(False)
            self.sync_info_label.setText("")

    def add_expression_graph(self):
        """Добавление результата вычисления как нового графика"""
        if not self.last_calculation_result or self.last_calculation_result['type'] != 'graph':
            return
        
        try:
            graph_data = self.last_calculation_result['graph_data']
            expression = self.last_calculation_result['expression']
            
            # Генерируем уникальный идентификатор
            expr_id = f"EXPR_{len(self.expression_lines)}"
            
            # Создаем график
            line, = self.ax.plot(graph_data.freqs, graph_data.amps, 
                                linewidth=2, linestyle='--', alpha=0.8)
            
            # Сохраняем данные
            key = ('Expression', expr_id, 'Result')
            self.frequency_responses[key] = (graph_data.freqs, graph_data.amps)
            self.lines[key] = line
            self.expression_lines[expr_id] = {
                'line': line,
                'expression': expression,
                'graph_data': graph_data,
                'freqs': graph_data.freqs,
                'amps': graph_data.amps
            }
            
            # Добавляем в легенду
            label = f"{graph_data.label}"
            if len(label) > 50:
                label = label[:50] + "..."
            color = line.get_color()
            self.legend_widget.add_line(label, line, color)
            
            # Обновляем масштаб
            if self.auto_update_cb.isChecked():
                self.auto_adjust_axes()
            
            self.canvas.draw()
            
            # Обновляем результат
            self.result_display.append(f"\n✓ График добавлен как '{label}'")
            
        except Exception as e:
            error_msg = f"Ошибка добавления графика: {str(e)}"
            logger.error(error_msg)
            self.result_display.append(f"\n{error_msg}")

    def reset_zoom(self):
        """Сброс масштаба к исходному"""
        if self.original_xlim and self.original_ylim:
            self.ax.set_xlim(self.original_xlim)
            self.ax.set_ylim(self.original_ylim)
            self.canvas.draw()

    def on_legend_visibility_changed(self):
        """Обработка изменения видимости через легенду"""
        if self.auto_update_cb.isChecked():
            self.auto_adjust_axes()
        self.canvas.draw()

    def load_selected_analyses(self):
        """Загрузка выбранных анализов и построение графиков в абсолютных величинах"""
        try:
            # Получаем выбранные анализы
            selected_analyses = self.tree_manager.get_selected_analyses()

            if not selected_analyses:
                QMessageBox.information(self, 'Информация', 'Не выбрано ни одного анализа')
                return

            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, len(selected_analyses))

            # Очищаем предыдущие данные (кроме выражений)
            keys_to_remove = []
            for key in list(self.frequency_responses.keys()):
                if key[0] != 'Expression':  # Сохраняем графики выражений
                    keys_to_remove.append(key)
            
            for key in keys_to_remove:
                if key in self.lines:
                    self.lines[key].remove()
                    del self.lines[key]
                del self.frequency_responses[key]
            
            # Очищаем легенду (кроме выражений)
            for label, checkbox in list(self.legend_widget.checkboxes.items()):
                if not label.startswith("(") and not label.startswith("ABS(") and not label.startswith("SQRT("):
                    self.legend_widget.checkbox_layout.removeWidget(checkbox)
                    checkbox.deleteLater()
                    del self.legend_widget.checkboxes[label]
                    if checkbox in self.legend_widget.lines_mapping:
                        del self.legend_widget.lines_mapping[checkbox]

            self.ax.clear()

            processed_count = 0
            valid_analyses = 0

            for i, (subject_code, analysis_index) in enumerate(selected_analyses):
                self.progress_bar.setValue(i)

                try:
                    # Получаем данные анализа
                    analysis_data = self.data_manager.get_analysis_data(subject_code, analysis_index)
                    if not analysis_data:
                        logger.warning(f"Данные анализа не найдены: {subject_code}, {analysis_index}")
                        continue

                    # Получаем процессор
                    processor = analysis_data.get('processor')
                    if not processor:
                        logger.warning(f"Процессор не найден: {subject_code}, {analysis_index}")
                        continue

                    # Получаем данные АЧХ в абсолютных величинах для всех каналов
                    freq_response_data = processor.freqresponse_linear

                    if not freq_response_data:
                        logger.warning(f"Нет данных АЧХ для анализа: {subject_code}, {analysis_index}")
                        continue

                    # Обрабатываем каждый канал
                    for channel_name, channel_data in freq_response_data.items():
                        freqs = channel_data['freq']
                        amplitude = channel_data['amplitude']  # Абсолютные величины в Вольтах

                        # Проверяем валидность данных
                        if (freqs is not None and amplitude is not None and
                            len(freqs) > 0 and len(amplitude) > 0 and
                            not np.all(np.isnan(amplitude)) and not np.all(np.isinf(amplitude))):

                            # Заменяем inf/nan на минимальное/максимальное значение
                            valid_mask = np.isfinite(amplitude)
                            if np.any(valid_mask):
                                # Используем только валидные точки
                                valid_freqs = freqs[valid_mask]
                                valid_amplitude = amplitude[valid_mask]

                                # Заменяем оставшиеся inf/nan
                                if len(valid_amplitude) > 0:
                                    min_amp = np.min(valid_amplitude)
                                    max_amp = np.max(valid_amplitude)
                                    valid_amplitude = np.where(
                                        np.isinf(valid_amplitude) | np.isnan(valid_amplitude),
                                        min_amp, valid_amplitude
                                    )

                                # Сохраняем данные
                                key = (subject_code, analysis_index, channel_name)
                                self.frequency_responses[key] = (valid_freqs, valid_amplitude)
                                valid_analyses += 1

                                # Строим график
                                label = f"{subject_code}_{analysis_index}_{channel_name}"
                                line, = self.ax.plot(valid_freqs, valid_amplitude, linewidth=2)
                                self.lines[key] = line

                                # Добавляем в легенду
                                color = line.get_color()
                                self.legend_widget.add_line(label, line, color)

                    processed_count += 1

                except Exception as e:
                    logger.error(f"Ошибка обработки анализа {subject_code}_{analysis_index}: {str(e)}")
                    continue

            # Добавляем обратно графики выражений
            for expr_id, expr_data in list(self.expression_lines.items()):
                try:
                    line = expr_data['line']
                    freqs = expr_data['freqs']
                    amps = expr_data['amps']
                    
                    # Перерисовываем линию
                    new_line, = self.ax.plot(freqs, amps, linewidth=2, linestyle='--', alpha=0.8)
                    
                    # Обновляем ссылки
                    key = ('Expression', expr_id, 'Result')
                    self.lines[key] = new_line
                    expr_data['line'] = new_line
                    
                    # Обновляем связь в легенде
                    for checkbox, old_line in list(self.legend_widget.lines_mapping.items()):
                        if old_line == line:
                            self.legend_widget.lines_mapping[checkbox] = new_line
                            break
                except Exception as e:
                    logger.error(f"Ошибка восстановления выражения {expr_id}: {str(e)}")
                    del self.expression_lines[expr_id]

            if valid_analyses == 0 and not self.expression_lines:
                self.ax.text(0.5, 0.5, 'Нет данных для построения графика',
                           transform=self.ax.transAxes, ha='center', va='center')
                self.ax.set_xlabel('Частота (Гц)')
                self.ax.set_ylabel('Амплитуда (В)')
                self.ax.set_title('Амплитудно-частотная характеристика')
            else:
                self.ax.set_xlabel('Частота (Гц)')
                self.ax.set_ylabel('Амплитуда (В)')
                self.ax.set_title(f'Сводный график АЧХ')
                self.ax.grid(True, alpha=0.3)

                # Сохраняем исходные пределы осей
                self.original_xlim = self.ax.get_xlim()
                self.original_ylim = self.ax.get_ylim()

                # Автоматически настраиваем масштаб
                self.auto_adjust_axes()

            self.canvas.draw()
            self.progress_bar.setVisible(False)

            logger.info(f"Обработано {processed_count} анализов, построено {valid_analyses} графиков")

            # Пересоздаем парсер с обновленными данными
            self.parser = ExpressionParser(self.frequency_responses)

        except Exception as e:
            logger.error(f"Ошибка при построении сводного графика: {str(e)}")
            QMessageBox.critical(self, 'Ошибка', f'Не удалось построить график: {str(e)}')
            self.progress_bar.setVisible(False)

    def auto_adjust_axes(self):
        """Автоматическая подстройка масштаба осей для видимых линий"""
        if not self.auto_update_cb.isChecked():
            return

        try:
            # Собираем все видимые данные
            all_visible_freqs = []
            all_visible_responses = []

            for key, line in self.lines.items():
                if line.get_visible():
                    if key in self.frequency_responses:
                        freqs, response = self.frequency_responses[key]
                        all_visible_freqs.extend(freqs)
                        all_visible_responses.extend(response)
                    elif key[0] == 'Expression':
                        # Для графиков выражений
                        expr_id = key[1]
                        if expr_id in self.expression_lines:
                            expr_data = self.expression_lines[expr_id]
                            all_visible_freqs.extend(expr_data['freqs'])
                            all_visible_responses.extend(expr_data['amps'])

            if not all_visible_freqs or not all_visible_responses:
                return

            # Преобразуем в numpy arrays для вычислений
            all_visible_freqs = np.array(all_visible_freqs)
            all_visible_responses = np.array(all_visible_responses)

            # Убираем NaN и inf значения
            valid_mask = np.isfinite(all_visible_responses) & np.isfinite(all_visible_freqs)
            if not np.any(valid_mask):
                return

            all_visible_freqs = all_visible_freqs[valid_mask]
            all_visible_responses = all_visible_responses[valid_mask]

            # Вычисляем пределы с небольшим отступом
            x_min, x_max = np.min(all_visible_freqs), np.max(all_visible_freqs)
            y_min, y_max = np.min(all_visible_responses), np.max(all_visible_responses)

            # Добавляем отступы (5% от диапазона)
            x_range = x_max - x_min
            y_range = y_max - y_min

            x_padding = x_range * 0.05 if x_range > 0 else 1
            y_padding = y_range * 0.05 if y_range > 0 else 1

            self.ax.set_xlim(x_min - x_padding, x_max + x_padding)
            self.ax.set_ylim(y_min - y_padding, y_max + y_padding)

        except Exception as e:
            logger.error(f"Ошибка при автоматической подстройке осей: {str(e)}")

    def get_visible_analyses(self):
        """Получение списка видимых анализов"""
        visible_keys = []
        for key, line in self.lines.items():
            if line.get_visible():
                visible_keys.append(key)
        return visible_keys

    def format_channel_parameters(self, channel_params, fixedlevel):
        """Форматирование параметров канала в текстовый вид"""
        if not channel_params:
            return "Параметры не рассчитаны"

        text = ""
        text += f"Максимальная амплитуда: {channel_params.get('max_amplitude', 0):.4f} В\n"
        text += f"Резонансная частота: {channel_params.get('resonance_frequency', 0):.2f} Гц\n"
        text += f"Ширина полосы (0.707): {channel_params.get('bandwidth_707', 0):.2f} Гц\n"

        bandwidth_707_range = channel_params.get('bandwidth_707_range', (0, 0))
        text += f"  (от {bandwidth_707_range[0]:.2f} до {bandwidth_707_range[1]:.2f} Гц)\n"

        text += f"Ширина полосы (уровень {fixedlevel}): {channel_params.get('bandwidth_fixed', 0):.2f} Гц\n"

        bandwidth_fixed_range = channel_params.get('bandwidth_fixed_range', (0, 0))
        text += f"  (от {bandwidth_fixed_range[0]:.2f} до {bandwidth_fixed_range[1]:.2f} Гц)\n"

        text += f"Добротность: {channel_params.get('q_factor', 0):.2f}"

        return text

    def export_to_excel(self):
        """Экспорт данных в Excel - только видимые графики в абсолютных величинах со всеми точками"""
        # Получаем только видимые анализы
        visible_keys = self.get_visible_analyses()

        if not visible_keys:
            QMessageBox.warning(self, 'Предупреждение', 'Нет видимых графиков для экспорта')
            return

        try:
            # Запрашиваем файл для сохранения
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                'Экспорт в Excel',
                'summary_analysis.xlsx',
                'Excel Files (*.xlsx)'
            )

            if not file_name:
                return

            self.progress_bar.setVisible(True)
            total_items = len(visible_keys)
            self.progress_bar.setRange(0, total_items + 2)

            # Создаем временное изображение графика
            temp_img_path = os.path.join(tempfile.gettempdir(), 'summary_plot.png')
            self.figure.savefig(temp_img_path, dpi=150, bbox_inches='tight')

            # Создаем Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Сводный анализ АЧХ"

            # Устанавливаем ширину колонок
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            # Заголовок
            title_cell = ws['A1']
            title_cell.value = "Сводный анализ АЧХ (только видимые графики, абсолютные величины)"
            title_cell.font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')
            title_cell.alignment = Alignment(horizontal='center')

            current_row = 3
            progress_count = 0

            # Определяем максимальное количество точек среди всех каналов
            max_points = 0
            all_data = {}  # Сохраняем все данные для второго прохода

            # Первый проход: собираем данные и находим максимальное количество точек
            for key in visible_keys:
                if key[0] == 'Expression':
                    # График выражения
                    expr_id = key[1]
                    if expr_id in self.expression_lines:
                        expr_data = self.expression_lines[expr_id]
                        freqs = expr_data['freqs']
                        amplitudes = expr_data['amps']
                        
                        if len(freqs) > 0 and len(amplitudes) > 0:
                            max_points = max(max_points, len(freqs))
                            all_data[key] = {
                                'freqs': freqs,
                                'amplitudes': amplitudes,
                                'label': f"Expression: {expr_data['expression']}",
                                'is_expression': True
                            }
                else:
                    subject_code, analysis_index, channel_name = key

                    analysis_data = self.data_manager.get_analysis_data(subject_code, analysis_index)
                    if not analysis_data:
                        continue

                    processor = analysis_data.get('processor')
                    if not processor:
                        continue

                    # Получаем параметры анализа
                    params = analysis_data.get('params', {})
                    if not isinstance(params, dict):
                        params = {}

                    # Получаем параметры канала
                    channel_params = {}
                    if hasattr(processor, 'channel_parameters'):
                        channel_params_dict = processor.channel_parameters
                        if isinstance(channel_params_dict, dict):
                            channel_params = channel_params_dict.get(channel_name, {})

                    # Получаем линейные данные АЧХ (абсолютные величины) - ВСЕ точки
                    linear_data = {}
                    if hasattr(processor, 'freqresponse_linear'):
                        linear_response_data = processor.freqresponse_linear
                        if channel_name in linear_response_data:
                            linear_data = linear_response_data[channel_name]

                    if linear_data and 'freq' in linear_data and 'amplitude' in linear_data:
                        freqs = linear_data['freq']
                        amplitudes = linear_data['amplitude']

                        if len(freqs) > 0 and len(amplitudes) > 0:
                            max_points = max(max_points, len(freqs))

                            # Сохраняем данные для второго прохода
                            all_data[key] = {
                                'freqs': freqs,
                                'amplitudes': amplitudes,
                                'subject_code': subject_code,
                                'analysis_index': analysis_index,
                                'channel_name': channel_name,
                                'params': params,
                                'channel_params': channel_params,
                                'is_expression': False
                            }

            # Второй проход: записываем данные в Excel по горизонтали
            current_col = 1  # Начинаем с колонки A (индекс 1)

            for key, data in all_data.items():
                progress_count += 1
                self.progress_bar.setValue(progress_count)

                freqs = data['freqs']
                amplitudes = data['amplitudes']
                
                if data['is_expression']:
                    # График выражения
                    label = data['label']
                    
                    # Заголовок выражения
                    title_cell = ws.cell(row=current_row, column=current_col)
                    title_cell.value = label
                    title_cell.font = Font(bold=True)
                    current_row += 1
                    
                    # Заголовки таблицы данных
                    freq_header = ws.cell(row=current_row, column=current_col)
                    freq_header.value = "Частота (Гц)"
                    freq_header.font = Font(bold=True)

                    amp_header = ws.cell(row=current_row, column=current_col + 1)
                    amp_header.value = "Амплитуда (В)"
                    amp_header.font = Font(bold=True)

                    current_row += 1
                else:
                    # Обычный график анализа
                    subject_code = data['subject_code']
                    analysis_index = data['analysis_index']
                    channel_name = data['channel_name']
                    params = data['params']
                    channel_params = data['channel_params']

                    # Заголовок анализа и канала
                    title_cell = ws.cell(row=current_row, column=current_col)
                    title_cell.value = f"Анализ: {subject_code}_{analysis_index} - Канал: {channel_name}"
                    title_cell.font = Font(bold=True)
                    current_row += 1

                    # Форматируем параметры канала
                    fixedlevel = params.get('fixedlevel', 0.6)
                    parameters_text = self.format_channel_parameters(channel_params, fixedlevel)

                    # Разбиваем текст параметров на строки и записываем в Excel
                    lines = parameters_text.split('\n')
                    for i, line in enumerate(lines):
                        param_cell = ws.cell(row=current_row + i, column=current_col)
                        param_cell.value = line

                    current_row += len(lines) + 1  # Отступ после параметров

                    # Заголовки таблицы данных
                    freq_header = ws.cell(row=current_row, column=current_col)
                    freq_header.value = "Частота (Гц)"
                    freq_header.font = Font(bold=True)

                    amp_header = ws.cell(row=current_row, column=current_col + 1)
                    amp_header.value = "Амплитуда (В)"
                    amp_header.font = Font(bold=True)

                    current_row += 1

                # Записываем ВСЕ точки данных
                for i in range(len(freqs)):
                    freq_cell = ws.cell(row=current_row + i, column=current_col)
                    amp_cell = ws.cell(row=current_row + i, column=current_col + 1)

                    freq_cell.value = float(freqs[i])
                    amp_cell.value = float(amplitudes[i])

                # Переходим к следующей группе колонок (с отступом в 2 колонки)
                current_col += 3
                # Сбрасываем строку для следующего канала
                current_row = 3

            # Добавляем изображение графика СПРАВА от данных
            if os.path.exists(temp_img_path):
                try:
                    img = XLImage(temp_img_path)
                    # Размещаем изображение справа от данных (колонка после последней группы данных)
                    image_start_col = current_col + 1
                    img.anchor = f'{chr(64 + image_start_col)}3'  # Например, 'E3' если current_col=4
                    ws.add_image(img)
                except Exception as e:
                    logger.error(f"Ошибка при добавлении изображения в Excel: {str(e)}")

            # Сохраняем файл
            try:
                wb.save(file_name)
                logger.info(f"Файл успешно сохранен: {file_name}")
            except Exception as e:
                logger.error(f"Ошибка при сохранении Excel файла: {str(e)}")
                raise

            # Удаляем временный файл
            try:
                os.remove(temp_img_path)
            except Exception as e:
                logger.warning(f"Не удалось удалить временный файл: {str(e)}")

            self.progress_bar.setVisible(False)
            QMessageBox.information(self, 'Успех', f'Данные экспортированы в {file_name}\n(только видимые графики, абсолютные величины, все точки)')

        except Exception as e:
            logger.error(f"Ошибка при экспорте в Excel: {str(e)}", exc_info=True)
            QMessageBox.critical(self, 'Ошибка', f'Не удалось экспортировать данные: {str(e)}')
            self.progress_bar.setVisible(False)